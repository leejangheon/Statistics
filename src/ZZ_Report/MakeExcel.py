#!/garnet2/Tools/META_Analysis/Tools/miniconda3/bin/python3
import os, sys
import pandas as pd
import json,subprocess

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


def parse_input(file, pvalue):
    sheetname = None
    title = None
    subtitle = None
    headers = []
    rows = []
    styles = []
    current_tag = None
    plist = []

    with open(file) as f:
        for line in f:
            line = line.rstrip("\n")
            if not line:
                continue

            if line.startswith("[") and line.endswith("]"):
                current_tag = line
                continue

            if current_tag == "[sheetname]":
                sheetname = line

            elif current_tag == "[Title]":
                title = line

            elif current_tag == "[SubTitle]":
                subtitle = line

            elif current_tag == "[th]":
                headers.append(line.split("\t"))

            elif current_tag == "[td]":
                rows.append(line.split("\t"))

            elif current_tag == "[style]":
                styles.append(line)

    with open(pvalue) as f:
        plist = [int(x) for x in f.read().strip().split(",")]

    return sheetname, title, subtitle, headers, rows, styles, plist



thin = Side(style="thin")

HEADER_FONT = Font(bold=True, size=10)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center")
HEADER_FILL = PatternFill(start_color="DCE6F1", fill_type="solid")
HEADER_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


DATA_LEFT_ALIGN = Alignment(horizontal="left", vertical="center")
DATA_CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

DATA_FONT = Font(size=10)
DATA_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

TITLE_FONT = Font(bold=True, size=16)
TITLE_ALIGN = Alignment(horizontal="center", vertical="center")
TITLE_FILL = PatternFill(start_color="DCE6F1", fill_type="solid")
TITLE_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

SUBTITLE_ALIGN = Alignment(vertical="center")


def merge_headers(ws, headers, start_row):

    nrow = len(headers)
    ncol = len(headers[0])

    # 가로 병합 
    for r in range(nrow):
        row = headers[r]
        c = 0
        while c < ncol:
            start = c
            val = row[c]

            if val == "":
                c += 1
                continue

            while c + 1 < ncol and row[c + 1] == val:
                c += 1

            end = c

            r1 = start_row + r + 1
            c1 = start + 1
            r2 = start_row + r + 1
            c2 = end + 1

            if start != end:
                ws.merge_cells(start_row=r1, start_column=c1,
                               end_row=r2, end_column=c2)
                
            cell = ws.cell(row=r1, column=c1, value=val)
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGN
            cell.fill = HEADER_FILL
            set_merged_border(ws, r1, c1, r2, c2)
            #cell.border = HEADER_BORDER

            c += 1

    # 세로 병합   
    for c in range(ncol):
        r = 0
        while r < nrow - 1:
            if headers[r][c] == headers[r + 1][c]:
                start = r
                val = headers[r][c]

                if val == "":
                    r += 1
                    continue

                while r + 1 < nrow and headers[r + 1][c] == val:
                    r += 1

                end = r

                r1 = start_row + start + 1
                r2 = start_row + end + 1
                c1 = c + 1

                ws.merge_cells(start_row=r1, start_column=c1,
                               end_row=r2, end_column=c1)
                set_merged_border(ws, r1, c1, r2, c1)
            r += 1


def set_merged_border(ws, r1, c1, r2, c2):
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):

            cell = ws.cell(row=row, column=col)
            b = cell.border  # 기존 border 가져오기

            cell.border = Border(
                left=thin if col == c1 else b.left,
                right=thin if col == c2 else b.right,
                top=thin if row == r1 else b.top,
                bottom=thin if row == r2 else b.bottom,
            )
def apply_styles(ws, styles):
    for s in styles:
        exec(s)

def convert_value(x):
    if x in ["None", ".",  "nan"]:
        return "."
    elif x == "":
        return ""
    try:
        return float(x)
    except:
        return x



def write_sheet(workbook, parsed):

    sheetname, title, subtitle, headers, rows, styles, plist = parsed
    ws = workbook.create_sheet(title=sheetname)

    headers_count = len(headers[0])

    # Title
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=headers_count)
    set_merged_border(ws, 1, 1, 1, headers_count)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = TITLE_ALIGN
    cell.fill = TITLE_FILL
    #cell.border = TITLE_BORDER

    ws.row_dimensions[1].height = 40

    # Subtitle
    ws.merge_cells(start_row=2, start_column=headers_count - 2,
                   end_row=2, end_column=headers_count)

    cell = ws.cell(row=2, column=headers_count - 2, value=subtitle)
    cell.alignment = SUBTITLE_ALIGN

    # Header
    start_row = 3
    merge_headers(ws, headers, start_row)

 
    df = pd.DataFrame(rows, columns=headers[-1])

    for i, r in df.iterrows():
        for j, v in enumerate(r):
            val = convert_value(v)
            if val=="":continue

            cell = ws.cell(
                row=start_row + len(headers) + i + 1,
                column=j + 1,
                value=val
            )

            # 스타일
            if isinstance(val, float):
                cell.alignment = DATA_CENTER_ALIGN
            else:
                cell.alignment = DATA_LEFT_ALIGN

            cell.border = DATA_BORDER
            cell.font = DATA_FONT
 
    max_row = len(df) + start_row + len(headers)
    apply_styles(ws,styles)
    for pcol in plist:
        col_letter = get_column_letter(pcol + 1)

        ws.conditional_formatting.add(
            f"{col_letter}{start_row+2}:{col_letter}{max_row}",
            CellIsRule(
                operator='lessThan',
                formula=['0.05'],
                font=Font(color="FF0000")
            )
        )

def front_page(wb,orderNumber,metadata,method,dirs):


    
    #1 get lims 
    # try >  , fail > Null 
    ws=wb['MACROGEN']
    api_url = f"https://lims03.macrogen.com/LIMSMOD/ngsDataMod2.jsp?on={orderNumber}"
    try:
       
        curl_output = subprocess.check_output(["curl", "-k", "-s", api_url])
        order_info = json.loads(curl_output)
        customer = order_info['customer']['name']
        institute = order_info['customer']['organization']
    except:
        print("[Waring] lims 정보 불러오기 에러 고객정보 none으로 표기됩니다.")
        print(f"[Waring] {api_url}")
        customer ="None"
        institute ="None"

    ws.cell(row=3, column=4, value=institute)
    ws.cell(row=4, column=4, value=customer)
    ws.cell(row=5, column=4, value=orderNumber)
    #2 get metadata
    df = pd.read_csv("metadata.txt", sep="\t")
    col = df.iloc[:, 1]
    col = col.astype(str).str.strip()
    total = len(col)

    counts = col.value_counts().sort_index()
    group_str = ", ".join([f"{g}:{c}" for g, c in counts.items()])

    result = f"{total} Samples ( {group_str} )"

   
    ws.cell(row=6, column=4, value=result)
    #3 dirs >> 

    txt_=""
    if method=="Kruskal":
        names = [os.path.basename(p) for p in dirs]
        analysis=", ".join(names)
        txt_=f"""{analysis} 에 대하여 Kruskal Waills Test를 진행
p-value <= 0.05인 항목에 대하여 dunn's post hoc 사후검정 결과 제공 

log2 Foldchange 계산식 :
         log2(A+ε)-log2(B+ε)
         •  ε (pseudo count) :  로그계산을 위해, 전체 데이터 중 0을 제외한 최소값보다 작은 보정 값"""

    elif method=="Wilcoxon":

        names = [os.path.basename(p) for p in dirs]
        analysis=", ".join(names)
        txt_=f"""{analysis} 에 대하여 Wilcoxon Rank Sum Test 를 진행

log2 Foldchange 계산식 :
         log2(A+ε)-log2(B+ε)
         •  ε (pseudo count) :  로그계산을 위해, 전체 데이터 중 0을 제외한 최소값보다 작은 보정 값"""

    elif method=="Wilcoxon_pair":

        names = [os.path.basename(p) for p in dirs]
        analysis=", ".join(names)
        txt_=f"""{analysis} 에 대하여 Wilcox Signed Rank Test 를 진행
"""

    ws.cell(row=7, column=4, value=txt_)

    
if __name__ == "__main__":

    orderNumber = sys.argv[1]
    metadata=sys.argv[2]
    method=sys.argv[3]
    dirs = sys.argv[4:]

    if len(dirs) == 0:
        print("Usage: python script.py dir1 dir2 ...")
        sys.exit(1)


    script_dir = os.path.dirname(os.path.abspath(__file__))
    if method == "Kruskal":
        templateFile = os.path.join(script_dir, "template", "Kruskal_Statistics_Result.xlsx")
    elif method == "Wilcoxon":
        templateFile = os.path.join(script_dir, "template", "WilcoxonRankSum_Statistics_Result.xlsx")
    elif method == "Wilcoxon_pair":
        templateFile = os.path.join(script_dir, "template", "WilcoxonSignedRank_Statistics_Result.xlsx")
    

    elif method == "t_test":
        templateFile = os.path.join(script_dir, "template", "t_test_Result.xlsx")

    elif method == "paired_t":
        templateFile = os.path.join(script_dir, "template", "Paired_t_test_Result.xlsx")

    elif method == "anova":
        templateFile = os.path.join(script_dir, "template", "ANOVA_Result.xlsx")
   
    else:
        templateFile = os.path.join(script_dir, "template", "Empty.xlsx")
    wb = load_workbook(templateFile)
    front_page(wb,orderNumber,metadata,method,dirs)


    #workbook = Workbook()
    # 기본 sheet 삭제
    #if "Sheet" in workbook.sheetnames:
    #   del workbook["Sheet"]


    for d in dirs:
        file = os.path.join(d, "final.output")
        pval = os.path.join(d, "pvalue.columns")

        parsed = parse_input(file, pval)
        write_sheet(wb, parsed)

    wb.save("output.xlsx")