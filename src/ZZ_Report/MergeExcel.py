#!/garnet2/Tools/META_Analysis/Tools/miniconda3/bin/python3
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import argparse
import os,sys
script_dir = os.path.dirname(os.path.abspath(__file__))

parser = argparse.ArgumentParser()
parser.add_argument("--order", type=str, default=None, help="header Name")
parser.add_argument("--method", type=str, default=None, help="header Name")
parser.add_argument("--sheet", type=str, default=None, help="sheet Name")

args = parser.parse_args()

# load & fill
# script dir / template /  #  Kruskal | wilcoxon | wilcoxon_pair | anova | t_test | paired_t_test 
if args.method =="Kruskal":
	templateFile = os.path.join(script_dir, "template", "Kruskal_Statistics_Result.xlsx")
	desc=""

elif args.method =="wilcoxon":
	templateFile = os.path.join(script_dir, "template", "WilcoxonRankSum_Statistics_Result.xlsx")
	desc=""

elif args.method =="wilcoxon_pair":
	templateFile = os.path.join(script_dir, "template", "WilcoxonSignedRank_Statistics_Result.xlsx")
	desc=""

wb = openpyxl.load_workbook(templateFile)
sheet = wb['MACROGEN']

#### main page 
## try  get lims
## fail >  custom , institute null
## Sample (per group Sample)




### subpage 
# for loop
# load and add


for f in files:
    sheet_wb = load_workbook(f)

    for sheet_name in sheet_wb.sheetnames:
        sheet_ws = sheet_wb[sheet_name]

        ws_new = wb.create_sheet(title=sheet_name)

        for row in ws.iter_rows(values_only=True):
            ws_new.append(row)